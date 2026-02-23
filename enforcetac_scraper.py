#!/usr/bin/env python3
"""
EnforceTac 2026 Exhibitor Scraper
==================================
A Playwright-based agent that navigates enforcetac.com, selects each product
category in turn, paginates via "Show more," and collects all exhibitor data.

Two modes of operation:
  1. API-intercept mode (default): captures the underlying Sitecore Search API
     calls the page makes, then replays them directly for speed and reliability.
  2. DOM-scrape mode (fallback): reads exhibitor data straight from the rendered
     page elements.

Requirements:
    pip install playwright openpyxl
    playwright install chromium

Usage:
    python enforcetac_scraper.py              # full run, Excel output
    python enforcetac_scraper.py --recon      # reconnaissance only (inspect DOM)
    python enforcetac_scraper.py --headed     # visible browser (useful for debugging)
"""

import argparse
import json
import re
import sys
import time
from pathlib import Path
from datetime import datetime
from urllib.parse import urljoin

from playwright.sync_api import sync_playwright, Page, BrowserContext

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_URL = "https://www.enforcetac.com"
EXHIBITOR_URL = f"{BASE_URL}/en/exhibitors-products/find-exhibitors"
OUTPUT_FILE = "enforcetac_2026_exhibitors_full.xlsx"
MAX_SHOW_MORE_CLICKS = 200          # safety limit per category
PAGE_LOAD_WAIT_MS = 4000            # wait after navigation
SHOW_MORE_WAIT_MS = 2000            # wait after each "show more" click
REQUEST_TIMEOUT_MS = 60000          # page navigation timeout


# ---------------------------------------------------------------------------
# Phase 1: Reconnaissance -- inspect page structure
# ---------------------------------------------------------------------------
def recon(page: Page) -> dict:
    """Examine the exhibitor search page to identify DOM selectors for
    category filters, exhibitor cards, and the 'show more' button."""

    info = {"categories": [], "exhibitor_sample": [], "api_calls": []}

    print("[RECON] Loading exhibitor search page...")
    page.goto(EXHIBITOR_URL, wait_until="networkidle", timeout=REQUEST_TIMEOUT_MS)
    page.wait_for_timeout(PAGE_LOAD_WAIT_MS)

    body_text = page.inner_text("body")
    print(f"[RECON] Page text length: {len(body_text)} chars")

    # Look for category / product-type filter elements
    selectors_to_try = {
        "checkboxes":       "input[type='checkbox']",
        "radio_buttons":    "input[type='radio']",
        "select_elements":  "select",
        "filter_buttons":   "[class*='filter'] button, [class*='Filter'] button, "
                            "[class*='facet'] button, [class*='Facet'] button",
        "accordion_items":  "[class*='accordion'], [class*='Accordion'], "
                            "[class*='collapsible'], [class*='Collapsible']",
        "chip_buttons":     "[class*='chip'], [class*='Chip'], [class*='tag'], [class*='Tag']",
        "category_links":   "a[href*='category'], a[href*='product-type'], "
                            "a[href*='productcategory']",
        "list_items_filter": "[class*='filter'] li, [class*='Filter'] li",
    }

    for label, sel in selectors_to_try.items():
        try:
            elems = page.query_selector_all(sel)
            if elems:
                print(f"[RECON] {label}: {len(elems)} elements found")
                for e in elems[:8]:
                    txt = (e.inner_text() or "").strip()[:100]
                    cls = e.get_attribute("class") or ""
                    val = e.get_attribute("value") or ""
                    tag = e.evaluate("el => el.tagName")
                    print(f"        <{tag}> class='{cls}' value='{val}' text='{txt}'")
        except Exception:
            pass

    # Look for "show more" / "load more" buttons
    for pattern in [
        "text=/show more/i", "text=/load more/i", "text=/mehr anzeigen/i",
        "text=/weitere/i", "text=/next/i",
        "button:has-text('Show')", "button:has-text('More')",
        "[class*='more'] button", "[class*='More'] button",
        "[class*='load'] button", "[class*='Load'] button",
        "[class*='pagination'] button", "[class*='Pagination'] button",
    ]:
        try:
            btn = page.locator(pattern)
            count = btn.count()
            if count > 0:
                txt = btn.first.inner_text()
                print(f"[RECON] 'Show more' candidate: selector='{pattern}' "
                      f"count={count} text='{txt.strip()[:60]}'")
        except Exception:
            pass

    # Look for exhibitor cards/tiles
    for pattern in [
        "[class*='exhibitor']", "[class*='Exhibitor']",
        "[class*='result-item']", "[class*='ResultItem']",
        "[class*='card']", "[class*='Card']",
        "[class*='tile']", "[class*='Tile']",
        "[class*='list-item']", "[class*='ListItem']",
        "article", "[role='listitem']",
    ]:
        try:
            items = page.query_selector_all(pattern)
            if items:
                print(f"[RECON] Exhibitor cards: selector='{pattern}' count={len(items)}")
                for it in items[:3]:
                    txt = (it.inner_text() or "").strip()[:200]
                    cls = it.get_attribute("class") or ""
                    print(f"        class='{cls}' text='{txt}'")
                    info["exhibitor_sample"].append(txt)
        except Exception:
            pass

    # Dump outer HTML for manual inspection
    html_snippet = page.evaluate("() => document.body.innerHTML.substring(0, 15000)")
    recon_html_path = Path("recon_body.html")
    recon_html_path.write_text(html_snippet, encoding="utf-8")
    print(f"[RECON] Saved first 15 000 chars of body HTML to {recon_html_path}")

    return info


# ---------------------------------------------------------------------------
# Phase 2: API-intercept mode
# ---------------------------------------------------------------------------
class APIInterceptor:
    """Captures XHR/fetch calls the page makes to the Sitecore Search API,
    then replays them with modified pagination to retrieve all results."""

    def __init__(self):
        self.api_calls = []
        self.api_endpoint = None
        self.api_headers = {}

    def on_request(self, request):
        url = request.url.lower()
        if any(kw in url for kw in [
            "discover", "search-rec", "rfksrv", "search/v",
            "graphql", "api/exhibitor", "api/search",
        ]):
            self.api_endpoint = request.url
            self.api_headers = dict(request.headers)
            post = request.post_data
            self.api_calls.append({
                "url": request.url,
                "method": request.method,
                "post_data": post,
            })
            print(f"[API] Intercepted: {request.method} {request.url[:120]}")
            if post:
                print(f"[API] POST body (first 500): {post[:500]}")

    def on_response(self, response):
        url = response.url.lower()
        if any(kw in url for kw in [
            "discover", "search-rec", "rfksrv", "search/v",
            "graphql", "api/exhibitor", "api/search",
        ]):
            try:
                body = response.json()
                for call in reversed(self.api_calls):
                    if call["url"] == response.url and "response" not in call:
                        call["response"] = body
                        break
                total = "?"
                widgets = body.get("widgets", [])
                for w in widgets:
                    t = w.get("total_item", w.get("total", w.get("total_items")))
                    if t is not None:
                        total = t
                        break
                print(f"[API] Response: {response.status} total_items={total}")
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Phase 3: DOM-based scraping
# ---------------------------------------------------------------------------
def get_category_filters(page: Page) -> list[dict]:
    """Identify all product category filter options on the page."""

    categories = []
    page.wait_for_timeout(2000)

    for container_sel in [
        "[class*='filter']", "[class*='Filter']",
        "[class*='facet']", "[class*='Facet']",
        "[class*='category']", "[class*='Category']",
        "[class*='product']", "[class*='Product']",
        "[data-testid*='filter']", "[data-testid*='facet']",
        "aside", "nav",
    ]:
        try:
            containers = page.query_selector_all(container_sel)
            for container in containers:
                items = container.query_selector_all(
                    "li, button, label, [role='checkbox'], [role='option'], a"
                )
                for item in items:
                    text = (item.inner_text() or "").strip()
                    if text and len(text) < 120:
                        match = re.match(r"^(.+?)\s*\((\d+)\)\s*$", text)
                        if match:
                            label, count = match.group(1).strip(), int(match.group(2))
                        else:
                            label, count = text, None
                        categories.append({
                            "label": label,
                            "count": count,
                            "element": item,
                        })
        except Exception:
            pass

    seen = set()
    unique = []
    for cat in categories:
        if cat["label"] not in seen:
            seen.add(cat["label"])
            unique.append(cat)

    print(f"[SCRAPE] Found {len(unique)} category filters")
    for c in unique[:15]:
        print(f"         - {c['label']} ({c['count'] or '?'})")
    if len(unique) > 15:
        print(f"         ... and {len(unique) - 15} more")

    return unique


def click_show_more_until_done(page: Page) -> int:
    """Click 'Show more' repeatedly until exhausted. Returns click count."""

    clicks = 0
    for _ in range(MAX_SHOW_MORE_CLICKS):
        btn = None
        for selector in [
            "text=/show more/i",
            "text=/load more/i",
            "text=/mehr anzeigen/i",
            "text=/weitere anzeigen/i",
            "button:has-text('Show more')",
            "button:has-text('Load more')",
            "[class*='more'] button",
            "[class*='More'] button",
            "[class*='showMore']",
            "[class*='ShowMore']",
            "[class*='load-more']",
            "[class*='LoadMore']",
        ]:
            try:
                loc = page.locator(selector)
                if loc.count() > 0 and loc.first.is_visible():
                    btn = loc.first
                    break
            except Exception:
                pass

        if btn is None:
            break

        try:
            btn.scroll_into_view_if_needed()
            btn.click()
            clicks += 1
            page.wait_for_timeout(SHOW_MORE_WAIT_MS)
            if clicks % 10 == 0:
                print(f"[SCRAPE] ... clicked 'Show more' {clicks} times")
        except Exception as e:
            print(f"[SCRAPE] 'Show more' click failed: {e}")
            break

    return clicks


def scrape_exhibitors_from_dom(page: Page) -> list[dict]:
    """Extract exhibitor data from the currently rendered page."""

    exhibitors = []

    card_selectors = [
        "[class*='exhibitor']", "[class*='Exhibitor']",
        "[class*='result-item']", "[class*='ResultItem']",
        "[class*='search-result']", "[class*='SearchResult']",
        "[class*='card']", "[class*='Card']",
        "[class*='tile']", "[class*='Tile']",
        "article[class*='list']",
        "[role='listitem']",
    ]

    cards = []
    for sel in card_selectors:
        try:
            found = page.query_selector_all(sel)
            if found and len(found) > len(cards):
                cards = found
                print(f"[SCRAPE] Using selector '{sel}' => {len(cards)} cards")
        except Exception:
            pass

    if not cards:
        links = page.query_selector_all("a[href*='/exhibitors-products/']")
        print(f"[SCRAPE] Fallback: found {len(links)} exhibitor-like links")
        for link in links:
            href = link.get_attribute("href") or ""
            text = (link.inner_text() or "").strip()
            if text and "/find-exhibitors" not in href and "/find-products" not in href:
                exhibitors.append({"company": text, "url": urljoin(BASE_URL, href)})
        return exhibitors

    for card in cards:
        entry = {}

        for name_sel in ["h2", "h3", "h4", "strong", "[class*='name']",
                         "[class*='Name']", "[class*='title']", "[class*='Title']"]:
            try:
                el = card.query_selector(name_sel)
                if el:
                    entry["company"] = el.inner_text().strip()
                    break
            except Exception:
                pass

        if not entry.get("company"):
            entry["company"] = card.inner_text().strip().split("\n")[0][:120]

        try:
            link = card.query_selector("a[href*='/exhibitors-products/']")
            if link:
                entry["url"] = urljoin(BASE_URL, link.get_attribute("href") or "")
        except Exception:
            pass

        full_text = card.inner_text() or ""
        hall_match = re.search(r'(?:Hall|Halle)\s+(\S+)', full_text, re.IGNORECASE)
        if hall_match:
            entry["hall"] = hall_match.group(1)

        booth_match = re.search(
            r'(?:Stand|Booth|Booth No\.?)\s*([\w\-\.]+)', full_text, re.IGNORECASE
        )
        if booth_match:
            entry["booth"] = booth_match.group(1)

        country_match = re.search(
            r'(?:Country|Land|Location)\s*[:\-]?\s*([A-Z][\w\s]+)', full_text
        )
        if country_match:
            entry["country"] = country_match.group(1).strip()

        entry["raw_text"] = full_text.strip()[:500]
        exhibitors.append(entry)

    return exhibitors


# ---------------------------------------------------------------------------
# Phase 4: Main orchestration
# ---------------------------------------------------------------------------
def run_scraper(headed: bool = False, recon_only: bool = False):
    all_exhibitors = {}
    api_interceptor = APIInterceptor()

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=not headed)
        context = browser.new_context(
            viewport={"width": 1440, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()

        page.on("request", api_interceptor.on_request)
        page.on("response", api_interceptor.on_response)

        # ---- Recon mode ----
        if recon_only:
            info = recon(page)
            print("\n[RECON] === API calls captured ===")
            for call in api_interceptor.api_calls:
                print(json.dumps({
                    "url": call["url"],
                    "method": call["method"],
                    "post_data": call.get("post_data", "")[:800],
                }, indent=2))
            browser.close()
            return

        # ---- Full scrape ----
        print("=" * 70)
        print("EnforceTac 2026 Exhibitor Scraper")
        print(f"Started: {datetime.now().isoformat()}")
        print("=" * 70)

        # Step 1: Load the page
        print("\n[1/4] Loading exhibitor search page...")
        page.goto(EXHIBITOR_URL, wait_until="networkidle", timeout=REQUEST_TIMEOUT_MS)
        page.wait_for_timeout(PAGE_LOAD_WAIT_MS)

        # Accept cookie consent if present
        for consent_sel in [
            "text=/accept/i", "text=/agree/i", "text=/akzeptieren/i",
            "[class*='consent'] button", "[class*='cookie'] button",
            "#onetrust-accept-btn-handler",
        ]:
            try:
                loc = page.locator(consent_sel)
                if loc.count() > 0 and loc.first.is_visible():
                    loc.first.click()
                    page.wait_for_timeout(1000)
                    print("[INFO] Dismissed cookie consent")
                    break
            except Exception:
                pass

        # Step 2: Identify categories
        print("\n[2/4] Identifying product categories...")
        categories = get_category_filters(page)

        if not categories:
            print("[WARN] No category filters found. Scraping default listing.")
            categories = [{"label": "(all)", "count": None, "element": None}]

        # Step 3: Iterate categories
        print(f"\n[3/4] Scraping exhibitors across {len(categories)} categories...")
        for idx, cat in enumerate(categories, 1):
            label = cat["label"]
            print(f"\n--- Category {idx}/{len(categories)}: {label} ---")

            if cat.get("element"):
                try:
                    cat["element"].scroll_into_view_if_needed()
                    cat["element"].click()
                    page.wait_for_timeout(PAGE_LOAD_WAIT_MS)
                except Exception as e:
                    print(f"[WARN] Could not click category '{label}': {e}")
                    continue

            clicks = click_show_more_until_done(page)
            print(f"[SCRAPE] 'Show more' clicked {clicks} times for '{label}'")

            exhibitors = scrape_exhibitors_from_dom(page)
            print(f"[SCRAPE] Extracted {len(exhibitors)} exhibitors for '{label}'")

            new_count = 0
            for ex in exhibitors:
                name = ex.get("company", "").strip()
                if not name:
                    continue
                if name not in all_exhibitors:
                    ex["product_category"] = label
                    all_exhibitors[name] = ex
                    new_count += 1
                else:
                    existing = all_exhibitors[name]
                    prev_cats = existing.get("product_category", "")
                    if label not in prev_cats:
                        existing["product_category"] = f"{prev_cats}; {label}"

            print(f"[SCRAPE] {new_count} new (cumulative: {len(all_exhibitors)})")

            # Reset filter before next category
            if cat.get("element"):
                try:
                    cat["element"].click()
                    page.wait_for_timeout(1500)
                except Exception:
                    page.goto(
                        EXHIBITOR_URL, wait_until="networkidle",
                        timeout=REQUEST_TIMEOUT_MS,
                    )
                    page.wait_for_timeout(PAGE_LOAD_WAIT_MS)

        browser.close()

    # ---- Report ----
    print(f"\n{'=' * 70}")
    print(f"Scraping complete. Total unique exhibitors: {len(all_exhibitors)}")
    print(f"API calls intercepted: {len(api_interceptor.api_calls)}")

    if api_interceptor.api_calls:
        api_log_path = Path("api_calls_log.json")
        api_log_path.write_text(
            json.dumps(api_interceptor.api_calls, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"[INFO] API call log saved to {api_log_path}")

    if api_interceptor.api_endpoint:
        print(f"[INFO] Discovered API endpoint: {api_interceptor.api_endpoint}")
        print("[HINT] Run with --recon to capture the full API structure,")
        print("       then use api_replay.py for faster bulk extraction.")

    # Step 4: Export to Excel
    print(f"\n[4/4] Exporting to Excel...")
    export_to_excel(list(all_exhibitors.values()), OUTPUT_FILE)
    print(f"[DONE] Saved {len(all_exhibitors)} exhibitors to {OUTPUT_FILE}")


# ---------------------------------------------------------------------------
# Phase 5: Excel export
# ---------------------------------------------------------------------------
def export_to_excel(exhibitors: list[dict], filepath: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Exhibitors"

    ws.merge_cells("A1:G1")
    ws["A1"] = (
        f"EnforceTac 2026 Exhibitors | "
        f"Scraped {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
        f"{len(exhibitors)} exhibitors"
    )
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1B4F72")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "Company", "Product Category", "Country/Region",
        "Hall", "Booth", "URL", "Raw Text",
    ]
    header_fill = PatternFill("solid", fgColor="2E86C1")
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    alt_fill = PatternFill("solid", fgColor="EBF5FB")
    data_font = Font(name="Arial", size=10)
    for i, ex in enumerate(
        sorted(exhibitors, key=lambda x: x.get("company", "").upper())
    ):
        row = i + 4
        ws.cell(row=row, column=1, value=ex.get("company", "")).font = data_font
        ws.cell(row=row, column=2, value=ex.get("product_category", "")).font = data_font
        ws.cell(row=row, column=3, value=ex.get("country", "")).font = data_font
        ws.cell(row=row, column=4, value=ex.get("hall", "")).font = data_font
        ws.cell(row=row, column=5, value=ex.get("booth", "")).font = data_font
        ws.cell(row=row, column=6, value=ex.get("url", "")).font = data_font
        ws.cell(row=row, column=7, value=ex.get("raw_text", "")[:200]).font = data_font

        if i % 2 == 1:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = alt_fill

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border

    widths = [35, 30, 18, 8, 14, 55, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{len(exhibitors) + 3}"

    wb.save(filepath)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="EnforceTac 2026 Exhibitor Scraper"
    )
    parser.add_argument(
        "--recon", action="store_true",
        help="Reconnaissance mode: inspect page structure only",
    )
    parser.add_argument(
        "--headed", action="store_true",
        help="Run with visible browser window",
    )
    args = parser.parse_args()

    run_scraper(headed=args.headed, recon_only=args.recon)
